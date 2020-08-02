#!/usr/bin/env python3.6
# -*- coding: utf-8 -*-
"""
Created on Thu Jun  4 13:57:45 2020

@author: Tom
"""
import os
import re
import cv2
import decimal
import numpy as np
import matplotlib.pyplot as plt

from model import Seg_unet
from trackmodel import Track_unet

from scipy import stats
from skimage import measure
from openpyxl import Workbook
from scipy.spatial.distance import euclidean
from scipy.ndimage.measurements import center_of_mass
from scipy.ndimage.morphology import binary_dilation as dilate
from matplotlib.widgets import Slider, Button, TextBox, CheckButtons
from scipy.spatial import ConvexHull, convex_hull_plot_2d


'''
These set of functions will run generate tracked segmentation masks by
combining the SegNet / TrackNet outputs, followed by a series of
analysis scripts.
'''

def ImportRawImages(imgpath = '../data/raw_images', img_type='tiff'):
    '''
    Function
    ----------
    ImportRawImages imports raw images in preparation for any analysis once
    tracking is completed that may need the image (i.e fluoresence)
    
    Use
    ----------
    Call this function as shown below:
    input_imgs = ImportRawImages()

    Parameters
    ----------
    path : directory
        path to raw_images directory within the folder template
    img_type : str
        default: tiff
        call as png if necessary.

    Returns
    -------
    inputs: list
        contains the processed images along with their filenames.
    
    
    '''
    # Get filenames from raw_images, and check for empty folder
    print('Importing timelapse images...')
    file_names = [name for name in os.listdir(imgpath) if img_type in name]
    assert(len(file_names) > 0), (('-->WARNING: No images of type {} found in '
                                  '-->raw_images').format(img_type))
    
    # Order the filenames by their frame number and check for completeness
    ord_names = sorted(file_names, key=FrameExtraction)
    frames = [FrameExtraction(name) for name in ord_names]
    missing = [frames[i] + 1 for i in range(len(frames) - 1)
               if frames[i + 1] != frames[i] + 1]
    if any(missing):
        print('-->WARNING: Frame(s) {} is/are missing'.format(missing))
        print('-->Use Ctrl+C to stop segmentation, or ignore and continue')
    else:
        print('No frames missing from raw_images import')
    
    # Import images
    raw_images = [(cv2.cvtColor(cv2.imread(os.path.join(imgpath, name)),
                   cv2.COLOR_BGR2GRAY)) for name in ord_names]
    print('{} timelapse frames successfully imported'.format(len(raw_images)))
    
    # Use ImageProcess to equalise, crop and resize each image
    print('Preparing images for segmentation...')
    input_imgs = [ImageProcess(img, raw=True) for img in raw_images] 
    print('Ready for segmentation')
    
    inputs = input_imgs,ord_names
    
    return inputs


def SegNet(inputs, imgpath='../data/raw_images/', snet_path='segnet.hdf5',
           binary_threshold=0.7):
    '''
    Function
    ----------
    SegNet() imports timelapse images from raw_images, passes them through the
    segmentation U-Net, and returns this raw network output along with binary
    segmentation masks.
    
    Use
    ----------
    Standard Use:
    masks = SegNet(input_imgs)
    
    Increased binary threshold :
    masks = SegNet(input_imgs,binary_threshold=0.8)
    
    Parameters
    ----------
    imgpath : path, optional
        The default is '../data/raw_images/'.
        Relative path to the raw_images folder from the current scripts dir
        
    snet_path : filename, optional
        The default is 'segnet.hdf5'.
        Name of the hdf5 file used to initalise the segnet.
        
        
    binary_threshold : float (between 0 and 1)
        The default is 0.7
        Increasing this threshold generates segmentation masks with fewer merge
        or false positive nuclei. Reducing it generates fewer false negative
        and split errors.

    Returns
    -------
    p_masks : array list
        raw network output, probabalistic masks [0,1].
        
    b_masks : array list
        binarised network output, binary masks [0 or 1]
        
    ord_names : string list
        list of ordered timelapse image filenames  (by frame)
    
    ** returned collectivley in the masks variable **
    '''
    input_imgs = inputs[0]
    ord_names = inputs[1]
    
    # Initialise the SegNet
    print('Initialising the SegNet...')
    segnet = Seg_unet(pretrained_weights=snet_path)
    
    # Generate predictions on each input image
    print('Generating segmentation masks...')
    p_masks = []
    for i, img in enumerate(input_imgs):
        p_mask = segnet.predict(input_imgs[i])
        p_masks.append(p_mask)
        print(p_mask.dtype)
        print('{} successfully segmented'.format(ord_names[i]))
    
    # Use Binarise to generate binary segmentation masks
    b_masks = [Binarise(mask, threshold=binary_threshold) for mask in p_masks]
    
    return (p_masks, b_masks, ord_names)

def AutoCorrectMasks(masks,binary_threshold=0.7):
    '''
    Parameters
    ----------
    masks : list of arrays
        Output from SegNet

    Returns
    -------
    None.

    '''
    # Unpack masks/names and label b_masks
    p_masks,b_masks,ord_names = masks
    labeled = [measure.label(np.squeeze(mask)) for mask in b_masks]
    updated = []
    
    # Cycle through each object in each frame
    for i,mask in enumerate(labeled):
        artefact = 0
        merge = 0
        for j in range(1,mask.max()+1):
            # Extract object
            obj = np.array(np.where(mask==j,1,0),dtype=np.uint8)
            
            # Remove if under-sized
            if np.sum(obj) < 30:
                mask = np.where(mask!=j,mask,0)
                artefact += 1
                continue
            
            # Get bounding box points
            xmax,ymax = [],[]
            for k in range(obj.shape[0]):
                if np.sum(obj[k,:]) > 0:
                    ymax.append(k)
                if np.sum(obj[:,k]) > 0:
                    xmax.append(k)
            area = (max(xmax)-min(xmax)) * (max(ymax)-min(ymax))
            ratio = np.sum(obj) / area  
            
            # Check for merged cells
            if ratio < 0.65:
                # Get p_mask object
                p_obj = np.squeeze(p_masks[i]) * obj
                
                # Re-binarse
                p_obj = np.where(p_obj>binary_threshold+0.15,1,0)
                
                # Only replace if binarisation successful
                if measure.label(p_obj).max() == 2:
                    # Remove old obj from mask
                    mask = np.where(mask != j,mask,0)
                    merge += 1
                    
                    # Add new binary object
                    mask += p_obj*20
        print('Frame {}: {} small artefact(s) and {} merge error(s) auto-corrected'.format(i,artefact,merge))
        updated.append(np.where(mask>0,1,0))
    
    # Save p_masks/b_masks to their respective folders
    print('Saving results to file...')
    for i in range(len(p_masks)):
        
        # Generate filename by removing TIFF and adding PNG
        p_name = os.path.splitext(ord_names[i])[0] + '.png'
        b_name = os.path.splitext(ord_names[i])[0] + '.png'
        p_name = os.path.join('../data/p_masks/', p_name)
        b_name = os.path.join('../data/b_masks/', b_name)
        
        # Get to file
        cv2.imwrite(p_name, np.squeeze(p_masks[i]) * 255)
        cv2.imwrite(b_name, np.squeeze(updated[i]) * 255)
    print('Auto-correction complete, masks saved to the b_masks folder in data')
        
                 
    return p_masks,updated,ord_names 

                               

def FrameExtraction(filename):
    '''
    Function
    ----------
    FrameExtraction extracts the frame number from a raw_image filename
    
    Use
    ---------
    This function is called automatically by SegNet

    Parameters
    ----------
    filename : string
        filename from raw_image

    Returns
    ---------
    frame : integer
        The integer frame number
        
    Errors:
    --------
    FRAMEKEY ERROR : Either the framekey variable requires updating because
    the labelling system for timelapse images has changed, or there
    are images incorrectly labeled in raw_images. To update the framekey
    variable, alter the frame key below to the new string of letters found
    immediatley before the frame number.
    '''
    # Assign framekey as string of characters before frame number in filename
    framekey = 'sk'  # <------ Update here if necessary
    assert(framekey in filename), ('FRAMEKEY ERROR: See FrameExtraction '
                                   'function of further instructions on '
                                   'updating the framekey')
    # Use regex to extract frame number
    regex = framekey + '(\d+)'
    frame = re.search(regex, filename).group(1)

    return int(frame)


def ImageProcess(i, raw=True):
    '''
    Function
    ----------
    ImageProcess equalises, crops and reshapes imports from raw_images.
    
    Use
    ----------
    This function is called automatically by SegNet()

    Parameters
    ----------
    i : np.array
        numpy array of size 1080x1080

    Returns
    ---------
    i : np.array
        numpy array of processed image now 1024x1024
    
    Errors
    ---------
    IMAGE SIZE ERROR: The images are of an incorrect shape. Raw_images must
    be 1080x1080, and any masks 1024x1024

    '''
    '''
    Process equalises, crops, resizes and converts raw images to float 32
    '''
    # Check for allowed image dimensions
    dim = (1080, 1080) if raw else (1024, 1024)
    assert(i.shape == dim), ('IMAGE SIZE ERROR: see '
                             'ImageProcess function '
                             'for further info')
    
    # Equlise, crop and reshape
    i = i / i.max()
    i = i[28:1052, 28:1052] if raw else i
    i = np.array(i, dtype=np.float32)
    i = np.reshape(i, (1, 1024, 1024, 1))
    
    return i


def Binarise(image, threshold=0.7):
    '''
    Function
    ----------
    Binarise binarises an image about a threshold.
    
    Use
    ----------
    This function is called automatically throughout analysis.py, but for
    manual binarisation call as follows:
    binary = Binarise(image)

    Parameters
    ----------
    image : np.array
        array of float values between 0 and 1.
        
    threshold : float, optional
        The default is 0.7. Anything aboveor equal to the threshold is set
        to 1. Anything below is set to 0.

    Returns
    -------
    binary : np.array (dtype = np.float)
        binary array of zeros and ones. The float dtype is required for
        assinging float labels later during tracking.

    '''
    # Use np.copy() to avoid altering orgional
    binary = np.copy(image)
    binary[binary < threshold] = 0.
    binary[binary >= threshold] = 1.
    
    return binary


def ViewMasks(masks):
    '''
    Function
    ---------
    ViewMasks plots all the binary segmentation masks to the plotting pane to
    be viewed.
    
    Use
    ---------
    Standard Use:
    ViewMasks(masks) where masks is the output from SegNet()
    
    Parameters
    ----------
    masks : list of np.arrays

    Returns
    -------
    None.

    '''
    # Unpack binary masks and filenames from the masks variable
    b_masks = masks[1]
    filenames = masks[2]
    
    # Set up a for loop for plotting each mask
    for i, mask in enumerate(b_masks):
        # Extract frame number from filename
        frame = FrameExtraction(filenames[i])
        
        # Generate plot
        plt.figure(dpi=300)  # Lower dpi if plotting very slow
        plt.imshow(np.squeeze(mask))
        plt.axis('off')
        plt.text(0, -10, 'Frame: ' + str(frame))  # Plots frame number
        plt.show()


def ImportMasks(b_path='../data/b_masks', p_path='../data/p_masks',
                img_type='png'):
    '''
    Function
    --------
    ImportMasks() imports binary and probabalistic masks that have been
    previously segmented
    
    Use
    --------
    masks = ImportMasks()
    
    Parameters
    --------
    b_path : path
        The default is ../data/b_masks
        Path to the binary mask folder
    p_path : path
        The default is ../data/p_masks
        Path to the probabalistic mask folder
    img_type : string, optional
        The default is 'png' (as saved by SegNet())
        The image type to search for in p/b masks. Can be updated to JPEG
        JPG, PNG, TIF, BMP. Must match the filename extension exactly.

    Returns
    --------
    masks : list of np.arrays
        In the same format as the SegNet() output, ready for tracking
    '''
    # Get filenames from p and b masks
    print('Importing masks...')
    pfile_names = [name for name in os.listdir(p_path) if img_type in name]
    assert(len(pfile_names) > 0), (('WARNING: No images of type {} found in '
                                    'p_masks').format(img_type))
    bfile_names = [name for name in os.listdir(b_path) if img_type in name]
    assert(len(bfile_names) > 0), (('WARNING: No images of type {} found in '
                                    'b_masks').format(img_type))
    # Order the filenames by their frame number
    p_names = sorted(pfile_names, key=FrameExtraction)
    b_names = sorted(bfile_names, key=FrameExtraction)
    
    # Get a list of frame numbers
    p_frames = [FrameExtraction(name) for name in p_names]
    b_frames = [FrameExtraction(name) for name in b_names]
    
    # Check all frames are present
    absent_p = [p_frames[i] + 1 for i in range(len(p_frames) - 1)
                if p_frames[i + 1] != p_frames[i] + 1]
    absent_b = [b_frames[i] + 1 for i in range(len(b_frames) - 1)
                if b_frames[i + 1] != b_frames[i] + 1]
    if any(absent_p):
        print('-->WARNING:Frame(s) {} is/are missing from p masks'.format(absent_p))
    if any(absent_b):
        print('-->WARNING:Frame(s) {} is/are missing from b masks'.format(absent_b))
    else:
        print('No frames are missing')
        
    # Import images
    p_masks = [(cv2.cvtColor(cv2.imread(os.path.join(p_path, name)),
                cv2.COLOR_BGR2GRAY)) for name in p_names]
    b_masks = [(cv2.cvtColor(cv2.imread(os.path.join(b_path, name)),
                cv2.COLOR_BGR2GRAY)) for name in b_names]
    
    # Use ImageProcess to equalise, crop and resize each image
    print('Processing masks')
    p_input = [ImageProcess(img, raw=False) for img in p_masks]
    b_input = [ImageProcess(img, raw=False) for img in b_masks]
    
    print('Import successful: {} masks imported'.format(len(p_input)))
    
    return p_input, b_input, p_names
    
      
def TrackNet(masks, tnet_path='tracknet.hdf5'):
    '''
    Function
    ----------
    TrackNet tracks nuclei across the binary segmentation masks generated by
    SegNet(). During tracking automatic quality control spots merge errors and
    attempts to fix them by altering the local binary threshold used to
    binarise the probabalistic network output. Any unsolveable errors are
    returned for later manual quality control
    
    Use
    ---------
    Standard Use:
    tracking_output = TrackNet(masks)
    
    Custom Range:
    tracking_output = TrackNet(masks,start=20,end=100)

    Parameters
    ----------
    masks : output from SegNet()
        triplet: p_masks, b_masks and filenames
        
    tnet_path : filename, optional
        The default is 'tracknet.hdf5'.
        Name of the hdf5 file used to initalise the tracknet.
    
    start/end : int
        If False, all timelapse frames will be tracked. Otherwise, assign the
        start and end frame to track over when calling TrackNet.

    Returns
    --------
    annot_frames : array list
        list of frame segmentations in which each nuclei has been assigned a
        unique ID.
        
    unsolveable_errors : dictionary
        stores any unsolveable errors for manual inspection later.
    
    Algorithm Overview
    ---------
    The following is an overview of a single round of trakcing. 'annot_frames'
    stores the labeled, seg masks in which each nucleus has a unique ID.
    Nuclei are pulled from the mask, with  their ID and run through the TrackNet
    along with whole current and future seg masks from the binary mask list.
    The output of tracking is used to build up the next annot_frame,
    nucleus by nucleus, until all nuclei have been added. Rather than use the
    tracking output directly, it is overlayed with the future b_mask to select
    the b_mask nucleus to which the current nucleus tracks to. This nucleus
    is then added to the growing annot_frame. Mitosis requires the addition
    of two nuclei, under new labels. Auto quality control is triggered when a
    merge error is detected in the future b_mask by tracking, the nuclei are
    then separated by tighetning the binary threshold in this area over the
    original p_mask. Frame number does not proceed to allow tracking to happen
    again with the corrected b_mask.
    '''

    # Unpack binary and probabalistic masks from masks
    masks = np.copy(masks)
    b_masks = [np.array(np.squeeze(mask),dtype=np.float32) for mask in masks[1]]
    p_masks = [np.array(np.squeeze(mask),dtype=np.float32) for mask in masks[0]]
    names = masks[2]
    
    # Initialise a dict to keep a record of any unsolveable issues
    unsolveable_errors = {}
        
    # Load tracknet
    print('Loading TrackNet')
    tracknet = Track_unet(input_size=(1024, 1024, 3))
    tracknet.load_weights(tnet_path)
    print('TrackNet loaded')
    
    # Label each nucleus in seed frame with unique integer ID
    annot_frames = [np.array(measure.label(b_masks[0]), dtype=np.float64)]
    
    # Iterate over frames
    frame = 0
    while frame < len(b_masks) - 1:
        print('Frame {} of {}...'.format(frame, len(b_masks)))
        
        # Initialise dictionary to store the labels from this round of tracking
        trackmap = {}
        
        # Get current and future segmentation masks
        curr_seg = np.reshape(b_masks[frame], (1024, 1024, 1))
        futr_seg = np.reshape(b_masks[frame + 1], (1024, 1024, 1))

        # Create empty array to which tracked nuclei will be added
        annot_frames.append(np.zeros((1024, 1024), dtype=np.float64))
        
        # Get current cell IDs from previous annot_frame
        IDs = np.unique(annot_frames[-2])[1:]  # Gets all non-zero labels
        
        # Iterate over each cell in frame
        for ID in IDs:
            # Isolate the cell being tracked, and prepare netowork inputs
            nucleus = np.reshape(np.where(annot_frames[-2] == ID, 1, 0), (1024, 1024, 1))
            inputs = np.concatenate(([nucleus], [curr_seg], [futr_seg]), axis=-1)
            
            # Predict, split and binarise outputs
            output = tracknet.predict(inputs)
            outputs = list(np.split(output, 3, axis=-1))[0:2]
            outputs = [np.squeeze(Binarise(array,threshold=0.5)) for array in outputs]
            
            # Rather than use the output of tracknet directly, overlap is used
            # get the actual nuclear footprint from the next seg mask
                        
            # Label the future segmentation
            lab_futr = measure.label(b_masks[frame + 1])
            
            # Get the label to which the cell in outputs[0] maps to
            overlap = lab_futr * outputs[0]
            futr_lab1 = stats.mode(np.where(overlap != 0, overlap, np.nan),
                                   axis=None, nan_policy='omit')[0][0]
        
            if futr_lab1 == 0:
                trackmap[ID] = 0
                continue
        
            # Check for mitosis
            if np.sum(outputs[1]) > 25:
                
                # Generate two new cell labels using custom nomenclature
                gen = -decimal.Decimal(str(ID)).as_tuple().exponent
                lab1 = round(ID + (1 / (10**(gen + 1))), gen + 1)
                lab2 = round(ID + (2 / (10**(gen + 1))), gen + 1)
                
                # Get the label to which the cell in outputs[1] maps to
                overlap = lab_futr * outputs[1]
                futr_lab2 = stats.mode(np.where(overlap != 0, overlap, np.nan),
                                       axis=None, nan_policy='omit')[0][0]
                
                # Check for false mitosis, by ensuring cell labels don't match
                if futr_lab1 != futr_lab2:
                    annot_frames[-1][lab_futr == futr_lab1] = lab1
                    annot_frames[-1][lab_futr == futr_lab2] = lab2
                    trackmap[ID] = (futr_lab1,futr_lab2)
                    
                elif futr_lab1 == futr_lab2:
                    annot_frames[-1][lab_futr == futr_lab1] = futr_lab1
                    trackmap[ID] = futr_lab1
                
            
            elif np.sum(outputs[0]) > 5:  # No mitosis, tracking still good
                # Add the determined nucleus from lab_futr to annot_frames
                # under the correct ID
                annot_frames[-1][lab_futr == futr_lab1] = ID
                trackmap[ID] = futr_lab1  # Also add mapping info to trackmap
                
        # Check for error events in the trackmap dict
        merge_IDs, absent_IDs = Check4Errors(trackmap)
                
        # If there are no errros, simply move onto next frame
        if not any(merge_IDs) and not any(absent_IDs):
            frame += 1
        # Otherwise, run auto-correction
        else:
            # Store the success of merge correction in success-log
            success_log = []
        
            # Deal with the merge type errors by tightening the binary threshold
            for IDa, IDb in merge_IDs:
                
                # Use the error IDs to get the merged nucleus from the most
                # recent annot_frame. It may either ID, so add both incase.
                merge_nuc = np.where(annot_frames[-1] == IDa, 1., 0.)
                merge_nuc += np.where(annot_frames[-1] == IDb, 1., 0.)
                
                # Use the m_nuc footprint to get the prob. network output nuc
                prob_nuc = p_masks[frame + 1] * merge_nuc
                
                # Increase the binary threshold over p_nuc until it splits
                for i in range(15):
                    threshold = 0.7 + 0.02 * i
                    new_nuc = measure.label(Binarise(prob_nuc, threshold=threshold))
                    
                    # Check for successful separation
                    if new_nuc.max() == 2:
                        success_log.append(True)
                        print('--> Merge error successfully auto-corrected')                  
                        # Replace the old merge_nuc in b_masks with the new one
                        b_masks[frame + 1] -= merge_nuc
                        b_masks[frame + 1] += np.where(new_nuc > 0, 1, 0)
                        
                        break # Break the binarising loop
                
                if new_nuc.max() != 2:
                    # If merge error not solveable, report and store in dict
                    success_log.append(False)
                    print('Auto-Quality Control failed on Frame {}, cells {}&{}'
                          .format(frame, IDa, IDb))
                    unsolveable_errors[frame] = (IDa, IDb)

            # Allow merge correction to cycle until all merges fixed and only
            # unsolveable errors remain (if at all), then correct absent errors
            if not any(success_log) or not True in success_log:
                
                # Iterate over any absent IDs
                for ID in absent_IDs:
                    absent_nuc = np.where(annot_frames[-2] == ID, 1, 0)
                    coords = center_of_mass(absent_nuc)
                    
                    # Get any nuclei that were unassigned by tracking
                    b_mask = np.where(b_masks[frame + 1] > 0, 1, 0)
                    annot_frame = np.where(annot_frames[frame + 1] > 0, 1, 0)
                    unmapped = measure.label(b_mask - annot_frame)
                    labels = np.unique(unmapped)[1:]
                    
                    # If there are no unassigned nuclei, correction impossible
                    if unmapped.max() == 0:
                        unsolveable_errors[frame] = ID
                        print('Auto-Quality Control failed on Frame {}, cell {}'
                              .format(frame, ID))
                        continue
    
                    # Store the coords of each nucleus under their labels
                    cdict = {label:center_of_mass(np.where(unmapped == label,1,0))
                             for label in labels}
                    
                    # Store the dist of each nuclei to absent_nuc under their labels 
                    dist = {label:euclidean(coords, cdict[label]) for label in [*cdict]}
                    
                    # Get the label of the nuclei closest to absent_nuc
                    sort = {k:v for k,v in sorted(dist.items(), 
                                                  key = lambda item:item[1])}
                    closest = [*sort][0]
                    
                    # Add to annot frames under the origional ID
                    annot_frames[-1] += np.where(b_masks[frame+1] == closest, ID, 0)
                    print('--> Tracking error successfully auto-corrected')
            
            # Return to previous frame and remove faulty frame if True in success
            if True in success_log:
                annot_frames = annot_frames[:-1]
            else:
                frame += 1

    return annot_frames, unsolveable_errors, names

def Check4Errors(trackmap):
    '''
    Function
    ----------
    Check4Errors checks a single round of tracking for any merge or absent
    type errors.
    
    Use
    ----------
    This function is called automatically in TrackNet()

    Parameters
    ----------
    trackmap : dictionary
        dict where future label of a nuclei is stored under its current ID

    Returns
    -------
    error_IDs : nested list
        list of pairs of merged IDs, as well as a list of absent nuclei in the
        final index
    '''
    # Remove any tuples from trackmap
    trackmap = {ID:trackmap[ID] for ID in [*trackmap] if type(trackmap[ID]) != tuple}

    # Get flat arrays of labels (Ls) and IDs
    Ls = np.array([trackmap[key] for key in [*trackmap]])
    IDs = np.array([*trackmap])
    
    # Get the counts of each label in Ls that is duplicated
    # Counts is a list of tuples, where each duplicated label is stored
    # alongside the number of times it appears in Ls.
    dups = [(la, np.sum(Ls == la)) for la in np.unique(Ls) if np.sum(Ls == la) == 2]
    
    # Now get the IDs that map to the duplicated labels
    merge_IDs = []
    for la, count in dups:
        if la == 0:  # Avoid a double absent error being counted as merge
            continue
        merge_IDs.append([ID for ID in IDs if trackmap[ID] == la])
        
    # Also add any IDs that map to nothing (absent error)
    absent_IDs = [[ID] for ID in IDs if trackmap[ID] == 0]
    
    return merge_IDs, absent_IDs


def SaveTrackingFrames(tracking_output, path='../data/tracking/'):
    
    frames = tracking_output[0]
    names = tracking_output[2]
    
    for i, img in enumerate(frames):
        
        img = np.where(img>0,img+10,img)
        
        plt.figure(dpi=200)
        plt.imshow(np.squeeze(img),cmap='gist_ncar')
        plt.axis('off')
        plt.savefig(path+names[i],dpi=200)
        plt.close()


coords=  []
def QualityControl(tracking_output): 
    '''
    QualityControl() allows the user to scroll through the annotated frames, and
    make manual error corrections
    '''    
    # Unpack annotated frames from tracking output
    frames = tracking_output[0]

    
    # Switch coloring of frames
    frames = [np.where(frame==0,frame.max()+5,frame) for frame in frames]
    
    # Plot the interactive figure starting at frame 0
    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(left=0.3, bottom=0.2)
    im1 = ax.imshow(frames[0])
    
    # Add each interactive axis
    axcolor = 'lightgoldenrodyellow'
    frameax = fig.add_axes([0.3, 0.1, 0.65, 0.03], facecolor=axcolor)
    nextax = fig.add_axes([0.6,0.01,0.07,0.07],facecolor=axcolor)
    backax = fig.add_axes([0.5,0.01,0.07,0.07],facecolor=axcolor)
    submitax = fig.add_axes([0.01,0.5,0.25,0.1],facecolor=axcolor)
    cancelax = fig.add_axes([0.01,0.4,0.25,0.1],facecolor=axcolor)
    swapframeax = fig.add_axes([0.01,0.3,0.25,0.1],facecolor=axcolor)
    
    # Define each slider/button/box object
    sframe = Slider(frameax, 'Frame', 0, 10, valinit=0,valstep=1,valfmt='%d')
    nextbutton = Button(nextax, 'Next', color=axcolor, hovercolor='0.975')
    backbutton = Button(backax, 'Back', color=axcolor, hovercolor='0.975')
    submitbutton = Button(submitax,'Submit',color=axcolor,hovercolor='0.975')
    cancelbutton = Button(cancelax,'Cancel',color=axcolor,hovercolor='0.975')
    swapframebox = CheckButtons(swapframeax,['Apply All'])
    
    # Set up Click Event
    def OnClick(event):
        if 'Subplot' not in str(event.inaxes):
            return
        ix, iy = event.xdata, event.ydata
        coords.append((int(iy),int(ix)))

        if len(coords) ==2:
            frame = int(sframe.val)
            cell1 = frames[frame][coords[0]]
            cell2 = frames[frame][coords[1]]
            fig.text(0.0,0.9,'Swap cells {} and {}?'
                     .format(cell1,cell2,frame))
            fig.canvas.draw_idle()
    fig.canvas.mpl_connect('button_press_event', OnClick)
    
    # Set up the slider update function
    def Update(val):
        im1.set_data(frames[int(val)])
        fig.canvas.draw_idle()
    sframe.on_changed(Update)
    
    # Set up the next button function
    def Next(event):
        frame = int(sframe.val+1)
        im1.set_data(frames[frame])
        sframe.set_val(frame)
        fig.canvas.draw_idle()
    nextbutton.on_clicked(Next)
    
    # Set up the back button function
    def Back(event):
        frame = int(sframe.val-1)
        im1.set_data(frames[frame])
        sframe.set_val(frame)
        fig.canvas.draw_idle()
    backbutton.on_clicked(Back)
    
    # Set up Cancel event
    def Cancel(event):
        coords.clear()
        fig.texts[0].remove()
    cancelbutton.on_clicked(Cancel)
    
    # Set up Submit event
    def Submit(event):
        # Get the necessary information to swap tracks
        frame = int(sframe.val)
        cell1 = frames[frame][coords[0]]
        cell2 = frames[frame][coords[1]]
        
        end = len(frames) if swapframebox.get_status()[0] else frame + 1
        
        # In this frame and all following, switch the IDs of the cells
        for i in range(frame,end):
            # Label one of the nuclei with a None label
            annot_frame = np.copy(frames[i])
            annot_frame = np.where(annot_frame==cell1,-1,annot_frame)

            # Now switch the label of the other nucleus
            annot_frame = np.where(annot_frame==cell2,cell1,annot_frame)

            # Now use the None label add the swaped label
            annot_frame = np.where(annot_frame==-1,cell2,annot_frame)

            # Re-assign the updated frame
            frames[i] = annot_frame
        fig.canvas.draw_idle()
    submitbutton.on_clicked(Submit)
    
    
    return sframe, nextbutton, backbutton, submitbutton, cancelbutton, swapframebox, coords, frames

global sel
sel = []

def TrackViewer(tracking_output):
    '''
    SelectTracks() allows the user to scroll through the annotated frames, for
    manual selection of cells across a given frame range. 
    '''    
    # Unpack annotated frames from tracking output
    frames = tracking_output[0]
    
    # Switch coloring of frames
    frames = [np.where(frame==0,frame.max()+5,frame) for frame in frames]
    
    # Plot the interactive figure starting at frame 0
    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(left=0.25, bottom=0.2)
    im1 = ax.imshow(frames[0])
    
    # Add each interactive axis
    axcolor = 'lightgoldenrodyellow'
    frameax = fig.add_axes([0.25, 0.1, 0.65, 0.03], facecolor=axcolor)
    nextax = fig.add_axes([0.6,0.01,0.07,0.07],facecolor=axcolor)
    backax = fig.add_axes([0.5,0.01,0.07,0.07],facecolor=axcolor)
    inputax = fig.add_axes([0.2,0.5,0.07,0.07], facecolor=axcolor)
    
    # Define each slider/button/box object
    sframe = Slider(frameax, 'Frame', 0, 10, valinit=0,valstep=1,valfmt='%d')
    nextbutton = Button(nextax, 'Next', color=axcolor, hovercolor='0.975')
    backbutton = Button(backax, 'Back', color=axcolor, hovercolor='0.975')
    inputbox = TextBox(inputax,'Trace Selection',label_pad=0.1)
    
    
    # Set up the slider update function
    def Update(val):
        im1.set_data(frames[int(val)])
        fig.canvas.draw_idle()
    sframe.on_changed(Update)
    
    # Set up the next button function
    def Next(event):
        frame = int(sframe.val+1)
        im1.set_data(frames[frame])
        sframe.set_val(frame)
        fig.canvas.draw_idle()
    nextbutton.on_clicked(Next)
    
    # Set up the back button function
    def Back(event):
        frame = int(sframe.val-1)
        im1.set_data(frames[frame])
        sframe.set_val(frame)
        fig.canvas.draw_idle()
    backbutton.on_clicked(Back)
    
    # Define the submission box
    sel = []
    def Submit(text):
        sel.append(text)
        inputbox.set_val(val='')
    inputbox.on_submit(Submit)
    
    return sframe,nextbutton,backbutton, inputbox, sel

def RecordSelection(text):
    global sel
    sel.append(text)
    return sel


def AnalyseAndExport(tracking_output,inputs,graphical_output):
    '''
    
    Function
    ----------
    Trace2XLS converts the annotated tracking frames into an excel datasheet
    It's at this stage the various metric extraction functions can be called 
    to extract the desired metric from the annotated tracking masks. 
    
    Use
    ----------
    Call this function as shown below, with the required inputs masks from
    SegNet and tracking_output from TrackNet:
    Trace2XLS(masks,tracking_output)

    Parameters
    ----------
    masks : list
        output of SegNet
    tracking_output : list
        output of TrackNet

    Returns
    -------
    None
    
    '''
    # Unpack inputs
    tracked_frames = tracking_output[0]
    images = [np.squeeze(img) for img in inputs[0]]
    selection = graphical_output[4]
    
    # Transform the selection list into a more useful form
    tracklist = {}
    selection = [sub for sub in selection if len(sub)>0]
    for sub in selection:
        # Get non '+' type submissions
        if '+' not in sub:
            # Non-tuple submissions need auto frame range finding
            sub = eval(sub)
            if type(sub) != tuple:
                ID = sub
                tracklist[ID] = Range(ID,tracked_frames)
            else:
                # Tuple submissions have a range associated with them
                tracklist[sub[0]] = (sub[1],sub[2])
        
        # Get '+' type submissions
        elif '+' in sub:
            # Get all IDs present across the tracked frames
            IDs = np.unique(tracked_frames)
            
            # Get non-tuple submissions
            if ',' not in sub:
                # Extract ID, find ancestors and append with their frame range
                IDx = eval(sub[:-1])
                tracklist[IDx] = Range(IDx,tracked_frames)
                
                for IDy in IDs:
                    if round(IDx) == round(IDy) and IDx != IDy:
                        tracklist[IDy] = Range(IDy,tracked_frames)
                
            else:
                # Tuple submissions have the range associated with them
                sub = sub.split(',')
                IDx = eval(sub[0][:-1])
                tracklist[IDx] = (eval(sub[1]),eval(sub[2]))
                for IDy in IDs:
                    if round(IDx) == round(IDy) and IDx != IDy:
                        tracklist[IDy] = Range(IDy,tracked_frames)
                
    print(tracklist)              
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tracking Results'
   
    # Add column titles
    ws['A1'] = 'Track ID'
    ws['B1'] = 'Frame'
    ws['C1'] = 'X center'
    ws['D1'] = 'Y center'
    ws['E1'] = 'Area'
    ws['F1'] = 'Nuclear Intensity'
    ws['G1'] = 'Ring Intensity'
    ws['H1'] = 'CDK2 Activity'

    # Iterate over the nuclei in tracklist, adding their metrics to the sheet
    j = 0
    for ID in tracklist:
        # Get the tracking range
        start,end = tracklist[ID]
        f_range = np.arange(start,end+1)
        
        # Get a list of frames/images with the isolated nuclei
        nuclei = [np.where(i==ID,1,0) for i in tracked_frames[start:end+1]]
        
        # Get the areas of the nucleus
        area = [np.sum(nucleus) for nucleus in nuclei]
        
        # Get the coordinate trace of the nucleus
        coords = [center_of_mass(nucleus) for nucleus in nuclei]
        X_coords = [coord[0] for coord in coords]
        Y_coords = [coord[1] for coord in coords]
        
        # Get the nuclear PCNA intensity from the images
        PCNA = [np.sum(i*nuc)/np.sum(nuc) for i in images[start:end+1]
                for nuc in nuclei]
        
        # For ring intensity, buffer nucleus with one pixel, then dilate +3
        buffer = [dilate(nuc,structure=np.ones((3,3))) for nuc in nuclei]
        dilated = [dilate(nuc,structure=np.ones((3,3)),iterations=4) for nuc in nuclei]
        
        # Generate rings by subtracting buffer from dilation
        rings = []
        for i,(buf,dil) in enumerate(zip(buffer,dilated)):
            rings.append(dil^buf)
            
        # Get ring intensity by multiplying ring by PCNA image
        I_ring = [np.sum(i*ring)/np.sum(ring) for i in images[start:end+1]
                  for ring in rings]
            
        # Add the above information to the spreadsheet
        for i,frame in enumerate(f_range):
            ws.cell(row=i+2+j,column=1,value=ID)
            ws.cell(row=i+2+j,column=2,value=frame)
            ws.cell(row=i+2+j,column=3,value=X_coords[i])
            ws.cell(row=i+2+j,column=4,value=Y_coords[i])
            ws.cell(row=i+2+j,column=5,value=area[i])
            ws.cell(row=i+2+j,column=6,value=PCNA[i])
            ws.cell(row=i+2+j,column=7,value=I_ring[i])
            ws.cell(row=i+2+j,column=8,value=I_ring[i]/PCNA[i])
            
        
        # Update the row offset
        j += end + 2 - start

    wb.save('Data.xlsx')



def Range(ID,tracked_frames):
    bools = [ID in frame for frame in tracked_frames]
    start = bools.index(True)
    end = start + sum(bools) - 1
    
    return start,end


